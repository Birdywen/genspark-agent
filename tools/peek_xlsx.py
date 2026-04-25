#!/usr/bin/env python3
import openpyxl

wb = openpyxl.load_workbook('/Users/yay/Downloads/article_review_lists.xlsx')
for name in wb.sheetnames:
    ws = wb[name]
    print(f'=== Sheet: {name} | Rows: {ws.max_row} | Cols: {ws.max_column} ===')
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(8, ws.max_row), values_only=False), 1):
        parts = []
        for cell in row:
            val = str(cell.value)[:80] if cell.value else ''
            link = ''
            if cell.hyperlink and cell.hyperlink.target:
                link = f' -> {cell.hyperlink.target[:100]}'
            parts.append(f'  [{cell.column_letter}] {val}{link}')
        print(f'Row {row_idx}:')
        for p in parts:
            print(p)
    print()
